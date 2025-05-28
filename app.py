import os
import uuid
import pandas as pd
from flask import Flask, render_template, request, send_file, session, redirect, url_for
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import xlwt 

app = Flask(__name__)
app.secret_key = 'clave-secreta-para-session'

CAMPAÑAS = [
    "BFBDELT1", "BFBT0A", "BFBT1A", "BFBT2A", "BFBT3A", "CNCBLOQ1",
    "CNCCAST1", "CNCT0", "CNCT1", "CRLEGAL", "FOHCASTI", "FOHT1A",
    "FOHVENC3", "GNBCAST1", "GNBVENC1", "IBK01", "IBK03", "IBK04", "IBK08",
    "IBKCAST1", "IBKT1A", "MAFCAST", "MAFJNR1", "MAFT2T3", "SANT1", "SANT2"
]

TEMP_FOLDER = 'temp'
TELEFONOS_FILE = 'telefonos.csv'

if not os.path.exists(TEMP_FOLDER):
    os.makedirs(TEMP_FOLDER)

def obtener_telefonos_por_campania(campania):
    df_tels = pd.read_csv(TELEFONOS_FILE)
    fila = df_tels[df_tels['campania'].str.upper() == campania.upper()]
    if fila.empty:
        return None
    return (
        fila.iloc[0]['phone_number_default'],
        fila.iloc[0]['phone_number_super'],
        fila.iloc[0]['phone_number_jefe']
    )

def guardar_xls(df, ruta):
    """Guardar un DataFrame en formato XLS con xlwt, y marcar filas (amarillo) para audio."""
    wb = xlwt.Workbook()
    ws = wb.add_sheet('Sheet1')

    # Estilo amarillo para fondo
    style_yellow = xlwt.easyxf('pattern: pattern solid, fore_colour yellow;')

    # Escribir encabezados
    for col_num, col_name in enumerate(df.columns):
        ws.write(0, col_num, col_name)

    # Escribir datos
    for row_num, row in enumerate(df.itertuples(index=False), start=1):
        for col_num, value in enumerate(row):
            # Para filas de audio insertadas (primeras 3 y últimas 3)
            if (row_num <= 3) or (row_num > len(df) - 3):
                ws.write(row_num, col_num, value, style_yellow)
            else:
                ws.write(row_num, col_num, value)

    wb.save(ruta)


@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        modulo = request.form.get('modulo')
        campaña = request.form.get('campaña')
        archivo = request.files.get('archivo')

        if modulo not in ['audio', 'mudo']:
            return "Módulo inválido", 400
        if campaña not in CAMPAÑAS:
            return "Campaña inválida", 400
        if not archivo:
            return "No se subió ningún archivo", 400

        # Leer Excel (xlsx o xls) con pandas (pandas usa xlrd para xls)
        df = pd.read_excel(archivo)
        df.iloc[:, 0] = df.iloc[:, 0].astype(str)
        df.iloc[:, 1] = df.iloc[:, 1].astype(str)
        df['city'] = campaña
        df['state'] = 1 if modulo == 'audio' else 0

        cantidad_ingresada = len(df)

        uid = str(uuid.uuid4())
        nombre_original = archivo.filename
        # Cambiamos extensión a .xls para el archivo temporal
        nombre_temp = f"{uid}_procesado.xls"
        ruta_temp = os.path.join(TEMP_FOLDER, nombre_temp)

        if modulo == 'audio':
            telefonos = obtener_telefonos_por_campania(campaña)
            if not telefonos:
                return f"No se encontraron teléfonos para la campaña {campaña}", 400
            tel_default, tel_super, tel_jefe = telefonos

            filas_insertar = pd.DataFrame([
                ["11111111", str(tel_default), campaña, 1],
                ["11111111", str(tel_super), campaña, 1],
                ["11111111", str(tel_jefe), campaña, 1]
            ], columns=df.columns)

            df_final = pd.concat([filas_insertar, df, filas_insertar], ignore_index=True)
            cantidad_final = len(df_final)

            # Guardar en XLS con xlwt para poder marcar las filas
            guardar_xls(df_final, ruta_temp)

        else:  # mudo
            cantidad_final = cantidad_ingresada
            # Guardar normal en xls sin filas extra ni color
            guardar_xls(df, ruta_temp)

        # Guardar info en session para descarga
        session['archivo_temporal'] = nombre_temp
        session['nombre_archivo'] = os.path.splitext(nombre_original)[0] + ".xls"
        session['cantidad_ingresada'] = cantidad_ingresada
        session['cantidad_final'] = cantidad_final
        session['modulo'] = modulo

        return render_template('indexGlobal.html',
                               campañas=CAMPAÑAS,
                               procesado=True,
                               nombre_archivo=session['nombre_archivo'],
                               cantidad_ingresada=cantidad_ingresada,
                               cantidad_final=cantidad_final,
                               modulo=modulo)

    return render_template('indexGlobal.html', campañas=CAMPAÑAS)

@app.route('/descargar')
def descargar():
    nombre_temp = session.get('archivo_temporal')
    nombre_original = session.get('nombre_archivo')

    if not nombre_temp or not nombre_original:
        return redirect(url_for('index'))

    ruta_temp = os.path.join(TEMP_FOLDER, nombre_temp)

    if not os.path.exists(ruta_temp):
        return "Archivo temporal no encontrado.", 404

    respuesta = send_file(ruta_temp,
                          as_attachment=True,
                          download_name=nombre_original,
                          mimetype='application/vnd.ms-excel')

    @respuesta.call_on_close
    def cleanup():
        try:
            os.remove(ruta_temp)
        except Exception as e:
            print(f"Error eliminando archivo temporal: {e}")

    return respuesta

if __name__ == '__main__':
    app.run(debug=True)
